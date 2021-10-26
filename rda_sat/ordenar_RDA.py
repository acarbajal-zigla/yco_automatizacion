import camelot
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries

import pandas as pd
import numpy as np

def get_headers_from_dataframe(df: pd.DataFrame):
    for index, row in df.iterrows():
        if (row[0] == "RFC") and (row[1] == "DENOMINACION SOCIAL"):
            alto_header = 1
            while(df.iloc[index+alto_header, 0] == "RFC"):
                alto_header += 1
            for column in range(2, len(row.tolist())):
                for i in range(1, alto_header):
                    row[column] = f"{row[column]} - {df.iloc[index+i, column]}"
            return [a.replace("\n", "").strip() for a in row.tolist()]
        else:
            continue

def set_categorias(df: pd.DataFrame):
    df.loc[df[df.columns[0]] == df[df.columns[1]], "CATEGORIA"] = df.loc[df[df.columns[0]] == df[df.columns[1]],df.columns[0]]
    df["CATEGORIA"] = df["CATEGORIA"].ffill()

def set_entidades_federativas(df: pd.DataFrame):
    df.loc[df[df.columns[0]].str.startswith("TOTAL "),"ENTIDAD FEDERATIVA"] = df.loc[df[df.columns[0]].str.startswith("TOTAL "),df.columns[0]]
    df.loc[:,"ENTIDAD FEDERATIVA"] = df.loc[:,"ENTIDAD FEDERATIVA"].str.replace("TOTAL ", "")
    df["ENTIDAD FEDERATIVA"] = df["ENTIDAD FEDERATIVA"].bfill()

def set_entidades_federativas_viejo(df: pd.DataFrame):
    df.loc[df["RFC"] == df[df.columns[2]],"ENTIDAD FEDERATIVA"] = df.loc[df["RFC"] == df[df.columns[2]],"RFC"]
    df["ENTIDAD FEDERATIVA"] = df["ENTIDAD FEDERATIVA"].ffill()


def get_header_row(ws:Worksheet):
    """
        Busca la fila en que aparecen los encabezados de la tabla.
        En caso de encontrarla, devuelve el número de fila (base 1)
        si no encuentra ningun header, devuelve False
    """
    for row in ws.iter_rows():
        if row[0].value == "ENTIDAD FEDERATIVA" or row[0].value == "RFC":
            print(row[0].row)
            return row[0].row
    return False

def get_headers(ws: Worksheet, row: int):
    headers = []
    alto_header = 1
    while ws.cell(row=row+alto_header, column=1).value == None:
        alto_header += 1
    row_data = row + alto_header
    
    i=1
    for columna in range(2, len(ws[row_data])+1):
        if ws.cell(row=row, column=columna+1).value == None:
            ws.cell(row=row, column=columna+1).value = ws.cell(row=row, column=columna).value

    i = 1
    while ws.cell(row=row_data+1, column=i).value:
        headers.append(ws.cell(row=row, column=i).value)
        print(headers)
        for row_header in range(1, alto_header):
            if ws.cell(row=row+row_header, column=i).value:
                headers[-1] += f" - {ws.cell(row=row+1, column=i).value}"
        i += 1

    return headers, row_data

def get_tabla_entidades_federativas(ws: Worksheet):
    header_row = get_header_row(ws)
    headers, first_data_row = get_headers(ws, header_row) 
    tabla_EF = pd.DataFrame(columns=headers)
    
    for row in ws.iter_rows(min_row=first_data_row):
        if row[0].value.startswith("TOTAL"):
            break
        else:
            tabla_EF.append([cell.value for cell in row])
    return tabla_EF

def get_table_from_pdf(path, hojas):

    tables = camelot.read_pdf(path, pages=hojas, copy_text=['h', 'v'],  line_scale=100)

    # Concateno todas las tablas en un dataframe
    df = pd.concat([t.df for t in tables])
    #df = df[df.columns[:-2]] # en los datos viejos elimina la columna de datos inválidos

    # Obtengo y asigno headers
    headers = get_headers_from_dataframe(df)
    #headers = [header for header in headers if (header.startswith("nan") == False)] # en datos viejos no contabiliza la columna nan-nan
    df.columns = headers

    # Elimino las filas que son repetición de headers por cambio de entidad federativa
    df = df[df["RFC"] != "RFC"]
    df = df[df["RFC"].str.startswith("TOTAL ") == False]

    # Elimino todas las filas vacias
    df = df.dropna()
    df = df[df["RFC"].astype(bool)]

    # Agrego las columnas de categoria y entidad federativa
    df["CATEGORIA"] = np.nan
    df["ENTIDAD FEDERATIVA"] = np.nan
    set_categorias(df)
    set_entidades_federativas(df)
    #set_entidades_federativas_viejo(df)

    df = df[df["RFC"] != df["CATEGORIA"]]
    df = df.replace("-","")
    return df